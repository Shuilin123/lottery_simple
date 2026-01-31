import os
import json
import shutil
import openpyxl

CACHE_DIR = os.path.join(os.path.dirname(__file__), 'cache')
os.makedirs(CACHE_DIR, exist_ok=True)

def load_temp_data():
    temp_path = os.path.join(CACHE_DIR, 'temp.json')
    error_path = os.path.join(CACHE_DIR, 'error.json')
    if os.path.exists(temp_path):
        with open(temp_path, 'r', encoding='utf-8') as f:
            temp = json.load(f)
    else:
        temp = {}
    if os.path.exists(error_path):
        with open(error_path, 'r', encoding='utf-8') as f:
            error = json.load(f)
    else:
        error = []
    return temp, error

def save_data_file(data):
    temp_path = os.path.join(CACHE_DIR, 'temp.json')
    with open(temp_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def save_error_data_file(data):
    error_path = os.path.join(CACHE_DIR, 'error.json')
    with open(error_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def shuffle(arr):
    import random
    random.shuffle(arr)

def load_excel(excel_path):
    if not os.path.exists(excel_path):
        return []
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        if data:
            data.pop(0)  # 去掉表头
        return [row for row in data if any(row)]
    except Exception:
        return []

def write_excel(data, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    out_path = os.path.join(os.getcwd(), filename)
    wb.save(out_path)
    return out_path
